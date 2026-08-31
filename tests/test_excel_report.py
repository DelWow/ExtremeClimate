import json
from dataclasses import replace
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import psycopg
import pytest
from openpyxl import load_workbook

from extreme_climate.excel_report import (
    ANOMALY_FILL_COLOR,
    ANOMALY_FONT_COLOR,
    CHART_TITLE,
    HEADER_FILL_COLOR,
    HEADER_FONT_COLOR,
    REPORT_HEADERS,
    WORKSHEET_TITLE,
    ReportGenerationError,
    ReportInputError,
    WeatherReportRow,
    build_weather_report_workbook,
    generate_weather_report,
    load_weather_report_rows,
    write_weather_report,
)


def _report_row(**changes) -> WeatherReportRow:
    row = WeatherReportRow(
        region_id="toronto",
        summary_date=date(2026, 8, 30),
        observation_count=24,
        mean_temperature_c=Decimal("21.25"),
        min_temperature_c=Decimal("16.50"),
        max_temperature_c=Decimal("26.75"),
        mean_humidity_percent=Decimal("61.50"),
        total_precipitation_mm=Decimal("2.25"),
        max_wind_speed_mps=Decimal("8.50"),
        is_anomaly=True,
        anomaly_status="anomaly",
        anomaly_details_json=(
            '{"anomalies":{"mean_temperature_c":{"deviation":5.25}},"status":"anomaly"}'
        ),
    )
    return replace(row, **changes)


def _database_row(
    *,
    region_id="toronto",
    summary_date=date(2026, 8, 30),
    is_anomaly=True,
    anomaly_details=None,
):
    if anomaly_details is None and is_anomaly is not None:
        anomaly_details = {
            "status": "anomaly" if is_anomaly else "normal",
            "anomalies": {},
        }
    return (
        region_id,
        summary_date,
        24,
        Decimal("21.25"),
        Decimal("16.50"),
        Decimal("26.75"),
        Decimal("61.50"),
        Decimal("2.25"),
        Decimal("8.50"),
        is_anomaly,
        anomaly_details,
    )


class FakeCursor:
    def __init__(self, rows):
        self.rows = tuple(rows)

    def fetchall(self):
        return self.rows


class FakeConnection:
    def __init__(self, *, rows=(), error=None):
        self.rows = tuple(rows)
        self.error = error
        self.execute_calls = []

    def execute(self, query, parameters):
        self.execute_calls.append((query, parameters))
        if self.error is not None:
            raise self.error
        return FakeCursor(self.rows)


def test_builds_one_sheet_with_fixed_headers_and_native_cell_types() -> None:
    workbook = build_weather_report_workbook([_report_row()])
    try:
        assert workbook.sheetnames == [WORKSHEET_TITLE]
        worksheet = workbook[WORKSHEET_TITLE]
        assert worksheet.max_row == 2
        assert worksheet.max_column == len(REPORT_HEADERS)
        assert tuple(cell.value for cell in worksheet[1]) == REPORT_HEADERS
        assert tuple(cell.value for cell in worksheet[2]) == (
            "toronto",
            date(2026, 8, 30),
            24,
            21.25,
            16.5,
            26.75,
            61.5,
            2.25,
            8.5,
            True,
            "anomaly",
            (
                '{"anomalies":{"mean_temperature_c":{"deviation":5.25}},'
                '"status":"anomaly"}'
            ),
        )
        assert len(worksheet._charts) == 1
    finally:
        workbook.close()


def test_rows_are_sorted_by_region_then_date() -> None:
    rows = [
        _report_row(summary_date=date(2026, 8, 31)),
        _report_row(region_id="halifax"),
        _report_row(summary_date=date(2026, 8, 29)),
    ]

    workbook = build_weather_report_workbook(rows)
    try:
        worksheet = workbook[WORKSHEET_TITLE]
        assert [
            (
                worksheet.cell(row=index, column=1).value,
                worksheet.cell(row=index, column=2).value,
            )
            for index in range(2, 5)
        ] == [
            ("halifax", date(2026, 8, 30)),
            ("toronto", date(2026, 8, 29)),
            ("toronto", date(2026, 8, 31)),
        ]
    finally:
        workbook.close()


def test_empty_input_produces_header_only_workbook() -> None:
    workbook = build_weather_report_workbook([])
    try:
        worksheet = workbook[WORKSHEET_TITLE]
        assert worksheet.max_row == 1
        assert tuple(cell.value for cell in worksheet[1]) == REPORT_HEADERS
        assert worksheet._charts == []
        assert list(worksheet.conditional_formatting) == []
    finally:
        workbook.close()


def test_applies_readable_header_and_data_formatting() -> None:
    workbook = build_weather_report_workbook([_report_row()])
    try:
        worksheet = workbook[WORKSHEET_TITLE]
        assert worksheet.freeze_panes == "A2"
        assert worksheet.auto_filter.ref == "A1:L2"
        assert worksheet.sheet_view.showGridLines is False
        assert worksheet.page_setup.orientation == "landscape"
        assert worksheet.print_title_rows == "$1:$1"

        assert worksheet["A1"].fill.fgColor.rgb == HEADER_FILL_COLOR
        assert worksheet["A1"].font.color.rgb == HEADER_FONT_COLOR
        assert worksheet["A1"].font.bold is True
        assert worksheet["A1"].alignment.wrap_text is True
        assert worksheet.row_dimensions[1].height == 34
        assert worksheet.column_dimensions["A"].width == 16
        assert worksheet.column_dimensions["D"].width == 23
        assert worksheet.column_dimensions["L"].width == 60

        assert worksheet["B2"].number_format == "yyyy-mm-dd"
        assert worksheet["C2"].number_format == "0"
        assert all(
            worksheet.cell(row=2, column=column).number_format == "0.00"
            for column in range(4, 10)
        )
        assert worksheet["L2"].alignment.wrap_text is True
    finally:
        workbook.close()


def test_adds_whole_row_anomaly_conditional_formatting() -> None:
    workbook = build_weather_report_workbook(
        [
            _report_row(is_anomaly=False, anomaly_status="normal"),
            _report_row(
                summary_date=date(2026, 8, 31),
                is_anomaly=True,
            ),
        ]
    )
    try:
        worksheet = workbook[WORKSHEET_TITLE]
        entries = list(worksheet.conditional_formatting)
        assert len(entries) == 1
        assert str(entries[0].sqref) == "A2:L3"
        rules = list(worksheet.conditional_formatting[entries[0]])
        assert len(rules) == 1
        rule = rules[0]
        assert rule.type == "expression"
        assert rule.formula == ["$J2=TRUE"]
        assert rule.stopIfTrue is True
        assert rule.dxf.fill.fgColor.rgb == ANOMALY_FILL_COLOR
        assert rule.dxf.font.color.rgb == ANOMALY_FONT_COLOR
    finally:
        workbook.close()


def test_temperature_chart_has_one_dated_series_per_region() -> None:
    rows = [
        _report_row(region_id="vancouver", summary_date=date(2026, 8, 31)),
        _report_row(region_id="toronto", summary_date=date(2026, 8, 31)),
        _report_row(region_id="halifax", mean_temperature_c=None),
        _report_row(region_id="vancouver"),
        _report_row(region_id="toronto"),
    ]

    workbook = build_weather_report_workbook(rows)
    try:
        worksheet = workbook[WORKSHEET_TITLE]
        assert len(worksheet._charts) == 1
        chart = worksheet._charts[0]
        assert chart.title.tx.rich.p[0].r[0].t == CHART_TITLE
        assert chart.anchor == "N2"
        assert chart.display_blanks == "gap"
        assert [series.tx.v for series in chart.series] == [
            "toronto",
            "vancouver",
        ]
        assert [series.val.numRef.f for series in chart.series] == [
            "'Daily Weather'!$D$3:$D$4",
            "'Daily Weather'!$D$5:$D$6",
        ]
        assert [series.cat.numRef.f for series in chart.series] == [
            "'Daily Weather'!$B$3:$B$4",
            "'Daily Weather'!$B$5:$B$6",
        ]
        assert all(series.marker.symbol == "circle" for series in chart.series)
    finally:
        workbook.close()


def test_omits_temperature_chart_when_all_temperature_values_are_missing() -> None:
    workbook = build_weather_report_workbook([_report_row(mean_temperature_c=None)])
    try:
        assert workbook[WORKSHEET_TITLE]._charts == []
    finally:
        workbook.close()


def test_optional_values_and_unevaluated_anomaly_remain_blank() -> None:
    workbook = build_weather_report_workbook(
        [
            _report_row(
                mean_humidity_percent=None,
                total_precipitation_mm=None,
                max_wind_speed_mps=None,
                is_anomaly=None,
                anomaly_status="not_evaluated",
                anomaly_details_json=None,
            )
        ]
    )
    try:
        row = tuple(cell.value for cell in workbook[WORKSHEET_TITLE][2])
        assert row[6:9] == (None, None, None)
        assert row[9:] == (None, "not_evaluated", None)
    finally:
        workbook.close()


def test_rejects_duplicate_region_date_rows() -> None:
    row = _report_row()

    with pytest.raises(ReportInputError, match="duplicate keys"):
        build_weather_report_workbook([row, row])


def test_loads_rows_with_canonical_anomaly_json() -> None:
    connection = FakeConnection(
        rows=[
            _database_row(
                anomaly_details={
                    "thresholds": {"temperature": 5},
                    "status": "normal",
                },
                is_anomaly=False,
            ),
            _database_row(
                region_id="vancouver",
                is_anomaly=None,
                anomaly_details=None,
            ),
        ]
    )

    rows = load_weather_report_rows(
        connection,
        date(2026, 8, 1),
        date(2026, 9, 1),
    )

    query, parameters = connection.execute_calls[0]
    assert "FROM weather_daily_summary" in query
    assert "ORDER BY region_id, summary_date" in query
    assert parameters == (date(2026, 8, 1), date(2026, 9, 1))
    assert rows[0].is_anomaly is False
    assert rows[0].anomaly_status == "normal"
    assert rows[0].anomaly_details_json == (
        '{"status":"normal","thresholds":{"temperature":5}}'
    )
    assert rows[1].is_anomaly is None
    assert rows[1].anomaly_status == "not_evaluated"
    assert rows[1].anomaly_details_json is None


def test_accepts_json_text_anomaly_details() -> None:
    connection = FakeConnection(
        rows=[
            _database_row(
                anomaly_details='{"status": "missing_baseline", "anomalies": {}}',
                is_anomaly=None,
            )
        ]
    )

    row = load_weather_report_rows(
        connection,
        date(2026, 8, 30),
        date(2026, 8, 31),
    )[0]

    assert row.anomaly_status == "missing_baseline"
    assert json.loads(row.anomaly_details_json)["anomalies"] == {}


@pytest.mark.parametrize(
    ("anomaly_details", "expected_error"),
    [
        ("not-json", "valid JSON"),
        ([], "JSON object"),
        ({"anomalies": {}}, "status must be"),
        ({"status": "normal", "value": float("nan")}, "JSON-compatible"),
    ],
)
def test_rejects_invalid_anomaly_details(
    anomaly_details,
    expected_error: str,
) -> None:
    connection = FakeConnection(rows=[_database_row(anomaly_details=anomaly_details)])

    with pytest.raises(ReportInputError, match=expected_error):
        load_weather_report_rows(
            connection,
            date(2026, 8, 30),
            date(2026, 8, 31),
        )


def test_writes_and_reopens_generated_workbook_programmatically(
    tmp_path: Path,
) -> None:
    destination = tmp_path / "nested" / "daily.xlsx"

    result = write_weather_report([_report_row()], destination)

    assert result.output_path == destination
    assert result.row_count == 1
    assert destination.is_file()
    workbook = load_workbook(destination, data_only=True)
    try:
        assert workbook.sheetnames == [WORKSHEET_TITLE]
        worksheet = workbook[WORKSHEET_TITLE]
        assert tuple(cell.value for cell in worksheet[1]) == REPORT_HEADERS
        assert worksheet["A2"].value == "toronto"
        assert worksheet["B2"].value.date() == date(2026, 8, 30)
        assert worksheet["C2"].value == 24
        assert worksheet["D2"].value == 21.25
        assert worksheet["J2"].value is True
        assert worksheet["K2"].value == "anomaly"
        assert worksheet["A1"].fill.fgColor.rgb == HEADER_FILL_COLOR
        assert len(worksheet._charts) == 1
        assert worksheet._charts[0].title.tx.rich.p[0].r[0].t == CHART_TITLE
        entries = list(worksheet.conditional_formatting)
        assert len(entries) == 1
        assert list(worksheet.conditional_formatting[entries[0]])[0].formula == [
            "$J2=TRUE"
        ]
    finally:
        workbook.close()


def test_generate_report_reads_then_saves_database_rows(tmp_path: Path) -> None:
    connection = FakeConnection(rows=[_database_row(is_anomaly=False)])
    destination = tmp_path / "report.xlsx"

    result = generate_weather_report(
        connection,
        date(2026, 8, 30),
        date(2026, 8, 31),
        destination,
    )

    assert result.row_count == 1
    workbook = load_workbook(destination, read_only=True, data_only=True)
    try:
        values = tuple(workbook[WORKSHEET_TITLE].values)
        assert values[1][0:3] == ("toronto", values[1][1], 24)
        assert values[1][9:11] == (False, "normal")
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("start_date", "end_date", "expected_error"),
    [
        (date(2026, 8, 30), date(2026, 8, 30), "must be before"),
        (date(2026, 8, 31), date(2026, 8, 30), "must be before"),
        (datetime(2026, 8, 30), date(2026, 8, 31), "must be dates"),
    ],
)
def test_rejects_invalid_report_date_range(
    start_date,
    end_date,
    expected_error: str,
) -> None:
    with pytest.raises(ReportInputError, match=expected_error):
        load_weather_report_rows(FakeConnection(), start_date, end_date)


def test_wraps_database_read_failure() -> None:
    connection = FakeConnection(error=psycopg.OperationalError("database unavailable"))

    with pytest.raises(ReportGenerationError, match="could not read"):
        load_weather_report_rows(
            connection,
            date(2026, 8, 30),
            date(2026, 8, 31),
        )


def test_rejects_non_xlsx_output_path(tmp_path: Path) -> None:
    with pytest.raises(ReportInputError, match="must end in .xlsx"):
        write_weather_report([], tmp_path / "report.csv")


def test_wraps_workbook_save_failure(tmp_path: Path) -> None:
    parent_file = tmp_path / "not-a-directory"
    parent_file.write_text("occupied", encoding="utf-8")

    with pytest.raises(ReportGenerationError, match="could not save"):
        write_weather_report([], parent_file / "report.xlsx")
