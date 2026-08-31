#!/usr/bin/env python3
"""Run and verify one deterministic local pipeline cycle."""

from __future__ import annotations

import json
import os
import subprocess
import sys
import time
import uuid
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Mapping, Sequence, Tuple

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src"))

import psycopg
import requests
from confluent_kafka import Consumer, KafkaException
from openpyxl import load_workbook

from extreme_climate.baseline_seed import (
    DEFAULT_BASELINE_FIXTURE_PATH,
    seed_historical_baselines,
)
from extreme_climate.kafka_publisher import (
    KafkaWeatherPublisher,
    load_kafka_settings,
)
from extreme_climate.postgres_config import load_postgres_settings
from extreme_climate.region_config import DEFAULT_REGIONS_PATH
from extreme_climate.weather_api import WeatherEvent
from extreme_climate.weather_consumer import (
    RawWeatherStore,
    load_weather_consumer_settings,
    process_message,
)


DAG_ID = "extreme_climate_pipeline"
E2E_REGION_ID = "toronto"
E2E_DATE = date(2099, 1, 15)
E2E_END_DATE = E2E_DATE + timedelta(days=1)
REPORT_PATH = PROJECT_ROOT / "reports" / (
    f"extreme_climate_daily_{E2E_DATE.isoformat()}.xlsx"
)
EXPECTED_TASK_IDS = (
    "validate_raw_weather",
    "transform_daily_weather",
    "detect_weather_anomalies",
    "generate_excel_report",
)


class EndToEndError(RuntimeError):
    """Raised when an end-to-end checkpoint does not match expectations."""


def _environment(**changes: str) -> Dict[str, str]:
    environment = dict(os.environ)
    environment.update(changes)
    return environment


def _ensure_stack() -> None:
    subprocess.run(
        ["docker", "compose", "up", "-d", "--wait"],
        cwd=PROJECT_ROOT,
        check=True,
    )


def _seed_baselines(environment: Mapping[str, str]) -> int:
    result = seed_historical_baselines(
        PROJECT_ROOT / DEFAULT_BASELINE_FIXTURE_PATH,
        PROJECT_ROOT / DEFAULT_REGIONS_PATH,
        load_postgres_settings(environment),
    )
    if result.stored_rows != 1830:
        raise EndToEndError(
            f"expected 1830 historical baseline rows, found {result.stored_rows}"
        )
    return result.stored_rows


def _event(run_marker: str) -> WeatherEvent:
    return WeatherEvent(
        region_id=E2E_REGION_ID,
        observed_at="2099-01-15T17:00:00Z",
        temperature_c=50.0,
        humidity_percent=20.0,
        precipitation_mm=25.0,
        wind_speed_mps=12.5,
        source_payload={
            "fixture": "step-17-end-to-end",
            "e2e_run_id": run_marker,
        },
    )


def _publish_event(environment: Mapping[str, str], event: WeatherEvent) -> int:
    settings = load_kafka_settings(environment)
    delivered = KafkaWeatherPublisher(settings).publish([event])
    if delivered != 1:
        raise EndToEndError(f"expected one Kafka delivery, got {delivered}")
    return delivered


def _consume_event(
    environment: Mapping[str, str],
    timeout_seconds: float = 30.0,
) -> Tuple[str, int, int]:
    settings = load_weather_consumer_settings(environment)
    consumer = Consumer(settings.consumer_config())
    connection = psycopg.connect(**settings.postgres.connection_kwargs())
    try:
        consumer.subscribe([settings.kafka.topic])
        deadline = time.monotonic() + timeout_seconds
        while time.monotonic() < deadline:
            try:
                message = consumer.poll(min(1.0, deadline - time.monotonic()))
            except KafkaException as exc:
                raise EndToEndError("Kafka polling failed") from exc
            if message is None:
                continue
            result = process_message(
                consumer,
                RawWeatherStore(connection),
                message,
                allowed_region_ids=settings.region_ids,
            )
            if not result.inserted:
                raise EndToEndError("the controlled Kafka event was a duplicate")
            return (
                result.position.topic,
                result.position.partition,
                result.position.offset,
            )
    finally:
        connection.close()
        consumer.close()
    raise EndToEndError("timed out waiting for the controlled Kafka event")


def _airflow_session(environment: Mapping[str, str]) -> requests.Session:
    session = requests.Session()
    session.auth = (
        environment.get("AIRFLOW_ADMIN_USERNAME", "airflow"),
        environment.get("AIRFLOW_ADMIN_PASSWORD", "airflow_dev"),
    )
    session.headers["Content-Type"] = "application/json"
    return session


def _airflow_url(environment: Mapping[str, str], path: str) -> str:
    port = environment.get("AIRFLOW_PORT", "8080")
    return f"http://127.0.0.1:{port}/api/v1{path}"


def _request_json(
    session: requests.Session,
    method: str,
    url: str,
    **kwargs: Any,
) -> Mapping[str, Any]:
    try:
        response = session.request(method, url, timeout=15, **kwargs)
        response.raise_for_status()
    except requests.RequestException as exc:
        detail = getattr(exc.response, "text", "")[:500]
        raise EndToEndError(f"Airflow API request failed: {detail}") from exc
    document = response.json()
    if not isinstance(document, Mapping):
        raise EndToEndError("Airflow API returned a non-object response")
    return document


def _run_airflow_dag(
    environment: Mapping[str, str],
    run_marker: str,
    timeout_seconds: float = 180.0,
) -> Tuple[str, Mapping[str, str]]:
    session = _airflow_session(environment)
    dag_url = _airflow_url(environment, f"/dags/{DAG_ID}")
    _request_json(session, "PATCH", dag_url, json={"is_paused": False})

    dag_run_id = f"e2e__{run_marker}"
    run_url = f"{dag_url}/dagRuns/{dag_run_id}"
    _request_json(
        session,
        "POST",
        f"{dag_url}/dagRuns",
        json={
            "dag_run_id": dag_run_id,
            "conf": {
                "window_start": E2E_DATE.isoformat(),
                "window_end": E2E_END_DATE.isoformat(),
            },
        },
    )

    deadline = time.monotonic() + timeout_seconds
    state = "queued"
    while time.monotonic() < deadline:
        run = _request_json(session, "GET", run_url)
        state = str(run.get("state"))
        if state in {"success", "failed"}:
            break
        time.sleep(2)
    if state != "success":
        raise EndToEndError(
            f"Airflow DagRun {dag_run_id} finished with state {state!r}"
        )

    task_document = _request_json(
        session,
        "GET",
        f"{run_url}/taskInstances",
    )
    task_instances = task_document.get("task_instances")
    if not isinstance(task_instances, Sequence):
        raise EndToEndError("Airflow did not return task instances")
    task_states = {
        str(task["task_id"]): str(task["state"])
        for task in task_instances
        if isinstance(task, Mapping)
    }
    if task_states != {task_id: "success" for task_id in EXPECTED_TASK_IDS}:
        raise EndToEndError(f"unexpected Airflow task states: {task_states}")
    return dag_run_id, task_states


def _database_results(
    environment: Mapping[str, str],
    run_marker: str,
) -> Tuple[Mapping[str, Any], Mapping[str, Any]]:
    settings = load_postgres_settings(environment)
    with psycopg.connect(**settings.connection_kwargs()) as connection:
        raw = connection.execute(
            """
            SELECT
                raw_weather_id,
                region_id,
                observed_at,
                kafka_topic,
                kafka_partition,
                kafka_offset
            FROM raw_weather
            WHERE source_payload ->> 'e2e_run_id' = %s
            """,
            (run_marker,),
        ).fetchone()
        summary = connection.execute(
            """
            SELECT
                region_id,
                summary_date,
                observation_count,
                mean_temperature_c,
                total_precipitation_mm,
                is_anomaly,
                anomaly_details
            FROM weather_daily_summary
            WHERE region_id = %s AND summary_date = %s
            """,
            (E2E_REGION_ID, E2E_DATE),
        ).fetchone()
    if raw is None:
        raise EndToEndError("controlled event is missing from raw_weather")
    if summary is None:
        raise EndToEndError("controlled date is missing from weather_daily_summary")
    raw_result = {
        "raw_weather_id": raw[0],
        "region_id": raw[1],
        "observed_at": raw[2].isoformat(),
        "kafka_topic": raw[3],
        "kafka_partition": raw[4],
        "kafka_offset": raw[5],
    }
    summary_result = {
        "region_id": summary[0],
        "summary_date": summary[1].isoformat(),
        "observation_count": summary[2],
        "mean_temperature_c": float(summary[3]),
        "total_precipitation_mm": float(summary[4]),
        "is_anomaly": summary[5],
        "anomaly_status": summary[6].get("status") if summary[6] else None,
    }
    if summary_result["is_anomaly"] is not True:
        raise EndToEndError(f"expected an anomalous summary: {summary_result}")
    if summary_result["anomaly_status"] != "anomaly":
        raise EndToEndError(f"unexpected anomaly details: {summary_result}")
    return raw_result, summary_result


def _verify_report(summary: Mapping[str, Any]) -> Mapping[str, Any]:
    if not REPORT_PATH.is_file():
        raise EndToEndError(f"expected report was not created: {REPORT_PATH}")
    workbook = load_workbook(REPORT_PATH, data_only=False)
    try:
        worksheet = workbook["Daily Weather"]
        matching_rows = [
            row
            for row in worksheet.iter_rows(min_row=2, values_only=True)
            if row[0] == E2E_REGION_ID and row[1].date() == E2E_DATE
        ]
        if len(matching_rows) != 1:
            raise EndToEndError(
                f"expected one report row for {E2E_REGION_ID}, found "
                f"{len(matching_rows)}"
            )
        row = matching_rows[0]
        report_result = {
            "region_id": row[0],
            "summary_date": row[1].date().isoformat(),
            "observation_count": row[2],
            "mean_temperature_c": row[3],
            "total_precipitation_mm": row[7],
            "is_anomaly": row[9],
            "anomaly_status": row[10],
        }
        expected = {
            key: summary[key]
            for key in report_result
        }
        if report_result != expected:
            raise EndToEndError(
                f"report row does not match PostgreSQL: {report_result}"
            )
        if not worksheet._charts:
            raise EndToEndError("report is missing its temperature chart")
        return report_result
    finally:
        workbook.close()


def run() -> Mapping[str, Any]:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    run_marker = f"{timestamp}_{uuid.uuid4().hex[:8]}"
    topic = f"weather_e2e_{run_marker.lower()}"
    environment = _environment(
        KAFKA_TOPIC=topic,
        KAFKA_CONSUMER_GROUP_ID=f"extreme-climate-e2e-{run_marker}",
        KAFKA_CONSUMER_CLIENT_ID=f"extreme-climate-e2e-{run_marker}",
        KAFKA_AUTO_OFFSET_RESET="earliest",
    )

    _ensure_stack()
    baseline_rows = _seed_baselines(environment)
    delivered = _publish_event(environment, _event(run_marker))
    kafka_position = _consume_event(environment)
    dag_run_id, task_states = _run_airflow_dag(environment, run_marker)
    raw_result, summary_result = _database_results(environment, run_marker)
    report_result = _verify_report(summary_result)

    return {
        "status": "passed",
        "run_id": run_marker,
        "baseline_rows": baseline_rows,
        "kafka": {
            "delivered": delivered,
            "topic": kafka_position[0],
            "partition": kafka_position[1],
            "offset": kafka_position[2],
        },
        "raw_weather": raw_result,
        "daily_summary": summary_result,
        "airflow": {
            "dag_run_id": dag_run_id,
            "task_states": task_states,
        },
        "report": {
            "path": str(REPORT_PATH.relative_to(PROJECT_ROOT)),
            "verified_row": report_result,
        },
    }


def main() -> int:
    try:
        result = run()
    except (EndToEndError, KafkaException, psycopg.Error) as exc:
        print(f"end-to-end verification failed: {exc}", file=sys.stderr)
        return 1
    except subprocess.CalledProcessError as exc:
        print(
            f"end-to-end verification failed: command exited {exc.returncode}",
            file=sys.stderr,
        )
        return 1
    print(json.dumps(result, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
