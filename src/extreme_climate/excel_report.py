"""Generate the formatted single-sheet Excel daily-weather report."""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable, Mapping, Optional, Protocol, Sequence, Tuple, Union

import psycopg
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.data_source import AxDataSource, NumRef
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_REPORT_PATH = Path("reports/extreme_climate_daily.xlsx")
WORKSHEET_TITLE = "Daily Weather"
CHART_TITLE = "Mean Temperature Trend"
CHART_ANCHOR = "N2"
HEADER_FILL_COLOR = "FF1F4E78"
HEADER_FONT_COLOR = "FFFFFFFF"
ANOMALY_FILL_COLOR = "FFFFC7CE"
ANOMALY_FONT_COLOR = "FF9C0006"
REPORT_HEADERS = (
    "Region ID",
    "Summary Date",
    "Observation Count",
    "Mean Temperature (°C)",
    "Minimum Temperature (°C)",
    "Maximum Temperature (°C)",
    "Mean Humidity (%)",
    "Total Precipitation (mm)",
    "Maximum Wind Speed (m/s)",
    "Is Anomaly",
    "Anomaly Status",
    "Anomaly Details",
)

_COLUMN_WIDTHS = (
    16,
    14,
    19,
    23,
    26,
    26,
    21,
    27,
    27,
    14,
    22,
    60,
)
_CHART_SERIES_COLORS = (
    "4472C4",
    "ED7D31",
    "70AD47",
    "A5A5A5",
    "FFC000",
    "5B9BD5",
)

_SELECT_REPORT_ROWS_SQL = """
    SELECT
        region_id,
        summary_date,
        observation_count,
        mean_temperature_c,
        min_temperature_c,
        max_temperature_c,
        mean_humidity_percent,
        total_precipitation_mm,
        max_wind_speed_mps,
        is_anomaly,
        anomaly_details
    FROM weather_daily_summary
    WHERE summary_date >= %s
      AND summary_date < %s
    ORDER BY region_id, summary_date
"""


class ReportInputError(ValueError):
    """Raised when report rows, dates, or anomaly details are malformed."""


class ReportGenerationError(RuntimeError):
    """Raised when report data cannot be read or the workbook cannot be saved."""


@dataclass(frozen=True)
class WeatherReportRow:
    """One typed row in the basic daily-weather workbook."""

    region_id: str
    summary_date: date
    observation_count: int
    mean_temperature_c: Optional[Decimal]
    min_temperature_c: Optional[Decimal]
    max_temperature_c: Optional[Decimal]
    mean_humidity_percent: Optional[Decimal]
    total_precipitation_mm: Optional[Decimal]
    max_wind_speed_mps: Optional[Decimal]
    is_anomaly: Optional[bool]
    anomaly_status: str
    anomaly_details_json: Optional[str]


@dataclass(frozen=True)
class ReportResult:
    """Location and row count of a saved workbook."""

    output_path: Path
    row_count: int


class CursorProtocol(Protocol):
    """Database cursor operations used while loading report rows."""

    def fetchall(self) -> Sequence[Sequence[Any]]: ...


class ConnectionProtocol(Protocol):
    """Database connection operations required by report generation."""

    def execute(self, query: str, params: Sequence[Any]) -> CursorProtocol: ...


def _validate_date_range(start_date: date, end_date: date) -> None:
    if (
        not isinstance(start_date, date)
        or isinstance(start_date, datetime)
        or not isinstance(end_date, date)
        or isinstance(end_date, datetime)
    ):
        raise ReportInputError("start_date and end_date must be dates")
    if start_date >= end_date:
        raise ReportInputError("start_date must be before end_date")


def _anomaly_fields(value: Any) -> Tuple[str, Optional[str]]:
    if value is None:
        return "not_evaluated", None
    if isinstance(value, str):
        try:
            document = json.loads(value)
        except json.JSONDecodeError as exc:
            raise ReportInputError("anomaly_details must be valid JSON") from exc
    else:
        document = value
    if not isinstance(document, Mapping):
        raise ReportInputError("anomaly_details must be a JSON object")
    status = document.get("status")
    if not isinstance(status, str) or not status:
        raise ReportInputError("anomaly_details.status must be a non-empty string")
    try:
        details_json = json.dumps(
            document,
            allow_nan=False,
            ensure_ascii=False,
            separators=(",", ":"),
            sort_keys=True,
        )
    except (TypeError, ValueError) as exc:
        raise ReportInputError("anomaly_details must be JSON-compatible") from exc
    return status, details_json


def _report_row(row: Sequence[Any]) -> WeatherReportRow:
    if len(row) != 11:
        raise ReportInputError("report query returned an invalid row")
    anomaly_status, anomaly_details_json = _anomaly_fields(row[10])
    return WeatherReportRow(
        region_id=row[0],
        summary_date=row[1],
        observation_count=row[2],
        mean_temperature_c=row[3],
        min_temperature_c=row[4],
        max_temperature_c=row[5],
        mean_humidity_percent=row[6],
        total_precipitation_mm=row[7],
        max_wind_speed_mps=row[8],
        is_anomaly=row[9],
        anomaly_status=anomaly_status,
        anomaly_details_json=anomaly_details_json,
    )


def load_weather_report_rows(
    connection: ConnectionProtocol,
    start_date: date,
    end_date: date,
) -> Tuple[WeatherReportRow, ...]:
    """Load report rows in the half-open date interval, sorted deterministically."""

    _validate_date_range(start_date, end_date)
    try:
        cursor = connection.execute(
            _SELECT_REPORT_ROWS_SQL,
            (start_date, end_date),
        )
        rows = cursor.fetchall()
    except psycopg.Error as exc:
        raise ReportGenerationError(
            "PostgreSQL could not read daily weather report data"
        ) from exc
    return tuple(_report_row(row) for row in rows)


def _excel_number(value: Optional[Decimal]) -> Optional[float]:
    return None if value is None else float(value)


def _format_worksheet(worksheet: Worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL_COLOR)
    header_font = Font(color=HEADER_FONT_COLOR, bold=True)
    thin_gray = Side(style="thin", color="FFD9E2F3")
    data_border = Border(bottom=thin_gray)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    worksheet.row_dimensions[1].height = 34

    for index, width in enumerate(_COLUMN_WIDTHS, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    for row_index in range(2, worksheet.max_row + 1):
        for cell in worksheet[row_index]:
            cell.border = data_border
            cell.alignment = Alignment(vertical="top")
        worksheet.cell(row=row_index, column=2).number_format = "yyyy-mm-dd"
        worksheet.cell(row=row_index, column=3).number_format = "0"
        for column_index in range(4, 10):
            worksheet.cell(row=row_index, column=column_index).number_format = "0.00"
        for column_index in (2, 3, 10, 11):
            worksheet.cell(row=row_index, column=column_index).alignment = Alignment(
                horizontal="center",
                vertical="top",
            )
        worksheet.cell(row=row_index, column=12).alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:L{worksheet.max_row}"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.print_title_rows = "1:1"

    if worksheet.max_row > 1:
        anomaly_fill = PatternFill(
            fill_type="solid",
            fgColor=ANOMALY_FILL_COLOR,
        )
        anomaly_font = Font(color=ANOMALY_FONT_COLOR)
        worksheet.conditional_formatting.add(
            f"A2:L{worksheet.max_row}",
            FormulaRule(
                formula=["$J2=TRUE"],
                fill=anomaly_fill,
                font=anomaly_font,
                stopIfTrue=True,
            ),
        )


def _add_temperature_chart(
    worksheet: Worksheet,
    rows: Sequence[WeatherReportRow],
) -> None:
    chart = LineChart()
    chart.title = CHART_TITLE
    chart.style = 13
    chart.y_axis.title = "Temperature (°C)"
    chart.x_axis.title = "Summary Date"
    chart.height = 9
    chart.width = 18
    chart.legend.position = "r"
    chart.display_blanks = "gap"

    start = 0
    color_index = 0
    while start < len(rows):
        region_id = rows[start].region_id
        end = start + 1
        while end < len(rows) and rows[end].region_id == region_id:
            end += 1
        region_rows = rows[start:end]
        if any(row.mean_temperature_c is not None for row in region_rows):
            first_excel_row = start + 2
            last_excel_row = end + 1
            values = Reference(
                worksheet,
                min_col=4,
                min_row=first_excel_row,
                max_row=last_excel_row,
            )
            categories = Reference(
                worksheet,
                min_col=2,
                min_row=first_excel_row,
                max_row=last_excel_row,
            )
            series = Series(values, title=region_id)
            series.cat = AxDataSource(numRef=NumRef(f=categories))
            series.marker.symbol = "circle"
            series.marker.size = 5
            series.graphicalProperties.line.solidFill = _CHART_SERIES_COLORS[
                color_index % len(_CHART_SERIES_COLORS)
            ]
            series.graphicalProperties.line.width = 24000
            chart.series.append(series)
            color_index += 1
        start = end

    if chart.series:
        worksheet.add_chart(chart, CHART_ANCHOR)


def build_weather_report_workbook(
    rows: Iterable[WeatherReportRow],
) -> Workbook:
    """Build a deterministic, formatted, single-sheet workbook."""

    report_rows = tuple(rows)
    keys = [(row.region_id, row.summary_date) for row in report_rows]
    if len(keys) != len(set(keys)):
        raise ReportInputError("weather report rows contain duplicate keys")
    report_rows = tuple(
        sorted(report_rows, key=lambda row: (row.region_id, row.summary_date))
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = WORKSHEET_TITLE
    worksheet.append(REPORT_HEADERS)
    for row in report_rows:
        worksheet.append(
            (
                row.region_id,
                row.summary_date,
                row.observation_count,
                _excel_number(row.mean_temperature_c),
                _excel_number(row.min_temperature_c),
                _excel_number(row.max_temperature_c),
                _excel_number(row.mean_humidity_percent),
                _excel_number(row.total_precipitation_mm),
                _excel_number(row.max_wind_speed_mps),
                row.is_anomaly,
                row.anomaly_status,
                row.anomaly_details_json,
            )
        )
    _format_worksheet(worksheet)
    _add_temperature_chart(worksheet, report_rows)
    return workbook


def write_weather_report(
    rows: Iterable[WeatherReportRow],
    output_path: Union[str, Path] = DEFAULT_REPORT_PATH,
) -> ReportResult:
    """Save report rows to an `.xlsx` file and return its path and row count."""

    destination = Path(output_path)
    if destination.suffix.lower() != ".xlsx":
        raise ReportInputError("output_path must end in .xlsx")
    report_rows = tuple(rows)
    workbook = build_weather_report_workbook(report_rows)
    try:
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
    except (OSError, ValueError) as exc:
        raise ReportGenerationError(
            f"could not save Excel report to {destination}"
        ) from exc
    finally:
        workbook.close()
    return ReportResult(output_path=destination, row_count=len(report_rows))


def generate_weather_report(
    connection: ConnectionProtocol,
    start_date: date,
    end_date: date,
    output_path: Union[str, Path] = DEFAULT_REPORT_PATH,
) -> ReportResult:
    """Read daily summaries and write their basic Excel report."""

    rows = load_weather_report_rows(connection, start_date, end_date)
    return write_weather_report(rows, output_path)
